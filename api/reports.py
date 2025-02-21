
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from .models import License

def report_as_excel(title, header, data, file_name):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    # Merge cells for title
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(header))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply column header styling
    for col_num, (column_title, column_width) in enumerate(header, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True, color='FFFFFF')
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = column_width
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(fill_type='solid', fgColor='01307A')

    # Populate data rows
    for row_num, row_data in enumerate(data, start=4):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(vertical='center')

    wb.save(response)
    return response